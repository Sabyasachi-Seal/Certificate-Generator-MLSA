import re
import os
import csv
import zipfile
import uvicorn
import subprocess
import aiofiles
import asyncio
from docx import Document
from typing import AsyncGenerator
from openpyxl import load_workbook
from fastapi.responses import HTMLResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from fastapi.middleware.cors import CORSMiddleware
from fastapi import FastAPI, Form, File, UploadFile, Request
from fastapi.responses import FileResponse
from certificate import (
    replace_participant_name,
    replace_event_name,
    replace_ambassador_name,
)
from send_mails import send_email
from threading import Thread, Semaphore
import sys
from typing import List
import time


class Email:
    def __init__(self, subject, email, html_content, attachment_path):
        self.subject = subject
        self.email = email
        self.html_content = html_content
        self.attachment_path = attachment_path

    def send(self):
        return asyncio.create_task(
            send_email(
                subject=self.subject,
                recipient=self.email,
                html_content=self.html_content,
                attachment_path=self.attachment_path,
            )
        )


app = FastAPI()

all_email_tasks: List[Email] = []

origins = ["*"]  # Adjust this to your frontend's actual origin(s)
app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

MAX_THREADS = 1

semaphore = Semaphore(MAX_THREADS)

# Serve static files (e.g., CSS, JS) from the 'static' folder
app.mount("/static", StaticFiles(directory="static"), name="static")

templates = Jinja2Templates(directory="templates")

mailerpath = "./Data/Mail.xlsm"
htmltemplatepath = "./Data/mailtemplate.html"
zip_filename = "./static/certificates.zip"
csv_file_path = "./Data/participants.csv"

# create output folder if not exist
try:
    os.makedirs("Output/Doc")
    os.makedirs("Output/PDF")
except OSError:
    pass


async def get_data_from_file() -> AsyncGenerator[bytes, None]:
    with open(file=zip_filename, mode="rb") as file_like:
        yield file_like.read()


def clear_mailer_file(wb, sheet):

    # Keep the first row (headers) and delete all other rows
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
        for cell in row:
            cell.value = None

    wb.save(filename=mailerpath)


def updatemailer(row, workbook, sheet, email, filepath, sub, body, status, cc=""):
    sheet.cell(row=row, column=1).value = email
    sheet.cell(row=row, column=2).value = cc
    sheet.cell(row=row, column=3).value = sub
    sheet.cell(row=row, column=4).value = body
    sheet.cell(row=row, column=5).value = filepath
    sheet.cell(row=row, column=6).value = status

    workbook.save(filename=mailerpath)


def getworkbook(filename):
    wb = load_workbook(filename=filename, read_only=False, keep_vba=True)
    sheet = wb.active
    return wb, sheet


def gethtmltemplate(htmltemplatepath=htmltemplatepath):
    return open(htmltemplatepath, "r").read()


def getmail(name, event, ambassador):
    sub = f"[MLSA] Certificate of Participation for {name}"
    html = gethtmltemplate(htmltemplatepath)
    body = html.format(name=name, event=event, ambassador=ambassador)
    return sub, body


async def process_csv(csv_file):

    participant_list = []

    first = True

    with open(csv_file) as csv_file:

        csv_reader = csv.reader(csv_file, delimiter=",")

        name_index = 0
        email_index = 0

        for line in csv_reader:
            if first:
                first = False
                for index, column in enumerate(line):
                    if re.search(r"\bName\b", column):
                        name_index = index
                    elif re.search(r"\bEmail\b", column):
                        email_index = index
                continue
            name = line[name_index]
            email = line[email_index]
            participant_list.append({"Name": name.strip(), "Email": email.strip()})

    return participant_list


def libreoffice_exec():
    # TODO: Provide support for more platforms
    if sys.platform == "darwin":
        return "/Applications/LibreOffice.app/Contents/MacOS/soffice"
    return "libreoffice"


class LibreOfficeError(Exception):
    def __init__(self, output):
        self.output = output


def convert_to_pdf(source, folder, timeout=None):

    args = [
        libreoffice_exec(),
        "--headless",
        "--convert-to",
        "pdf",
        "--outdir",
        folder,
        source,
    ]

    process = subprocess.run(
        args, stdout=subprocess.PIPE, stderr=subprocess.PIPE, timeout=timeout
    )
    print(process.stdout, process.stderr)


def convert_to_pdf_old(input_path, output_path):
    cmd = ["unoconv", "-f", "pdf", "-o", output_path, input_path]
    process = subprocess.Popen(cmd, stderr=subprocess.PIPE)
    output, error = process.communicate()
    print(output, error)
    return False, error


def zip_folder(folder_path, zip_filename, additional_files):

    # print(os.listdir(folder_path))

    with zipfile.ZipFile(zip_filename, "w", zipfile.ZIP_DEFLATED) as zipf:
        for root, dirs, files in os.walk(folder_path):
            for file in files:
                file_path = os.path.join(root, file)
                arcname = os.path.relpath(file_path, folder_path)
                zipf.write(file_path, arcname)

        # Include additional files
        if additional_files:
            for file_path in additional_files:
                arcname = os.path.relpath(file_path, os.path.dirname(file_path))
                zipf.write(file_path, arcname)


def worker(index, participate, wb, sheet, event, ambassador, filename):
    # Acquire a semaphore
    semaphore.acquire()
    try:
        make_certificates(index, participate, wb, sheet, event, ambassador, filename)
    finally:
        # Release the semaphore
        semaphore.release()


def make_certificates(index, participate, wb, sheet, event, ambassador, filename):
    print(participate)

    name = participate["Name"]
    email = participate["Email"]

    if email == "" or name == "":
        return

    doc = Document(filename)

    replace_participant_name(doc, name)
    replace_event_name(doc, event)
    replace_ambassador_name(doc, ambassador)

    doc.save("./Output/Doc/{}.docx".format(name))

    convert_to_pdf(
        "./Output/Doc/{}.docx".format(name),
        "./Output/PDF/",
    )

    filepath = os.path.abspath("./Output/PDF/{}.pdf".format(name))

    sub, body = getmail(name, event, ambassador)

    if wb and sheet:

        updatemailer(
            row=index + 2,
            workbook=wb,
            sheet=sheet,
            email=email,
            filepath=filepath,
            sub=sub,
            body=body,
            status="Send",
        )

    all_email_tasks.append(
        Email(
            subject=sub,
            email=email,
            html_content=body,
            attachment_path=filepath,
        )
    )


async def create_docx_files(filename, list_participate, event, ambassador):

    try:
        wb, sheet = getworkbook(mailerpath)
    except Exception as e:
        print(e)
        wb, sheet = None, None

    clear_mailer_file(wb, sheet) if wb and sheet else None

    os.system("rm -rf Output/Doc/*")
    os.system("rm -rf Output/PDF/*")

    # List of all tasks
    tasks = list(enumerate(list_participate))

    while tasks:
        threads = []
        # Start MAX_THREADS threads
        for _ in range(min(MAX_THREADS, len(tasks))):
            index, participate = tasks.pop(0)
            t = Thread(
                target=worker,
                args=(index, participate, wb, sheet, event, ambassador, filename),
            )
            t.start()
            threads.append(t)

        # Wait for these threads to finish
        for t in threads:
            t.join()


@app.head("/")
def read_root_head():
    return {"message": "MLSA Certificate Generator is running"}


@app.get("/", response_class=HTMLResponse)
def read_item(request: Request):
    return templates.TemplateResponse("index.html", context={"request": request})


@app.post("/send_emails")
async def send_emails():
    global all_email_tasks
    for email_task in all_email_tasks:
        try:
            task = email_task.send()
            results = await asyncio.gather(task, return_exceptions=True)
            if isinstance(results[0], Exception):
                raise results[0]
            print(f"Email sent to {email_task.email}")
        except Exception as e:
            print(f"Error sending email to {email_task.email}: {e}")
            continue

        time.sleep(2)

    # Clear the tasks list after processing
    all_email_tasks.clear()

    return {"message": "Emails sent successfully"}


@app.get("/{filepath}")
def get_file(filepath: str):
    file_path = os.path.abspath(os.path.join("./static", filepath))
    # print(file_path)
    return FileResponse(file_path)


async def get_statinfo():
    with open(zip_filename, "rb") as file:
        yield file.read()


def is_valid_csv(file_content: str) -> bool:
    try:
        # Attempt to parse the CSV content
        csv.reader(file_content.splitlines())
        return True
    except csv.Error:
        return False


@app.post("/generate_certificates")
async def generate_certificates(
    event_name: str = Form(...),
    ambassador_name: str = Form(...),
    participant_file: UploadFile = File(...),
):

    async with aiofiles.open(csv_file_path, "wb") as buffer:

        # Read the content of the uploaded file
        content = await participant_file.read()

        # Check if the content is a valid CSV
        if not is_valid_csv(content.decode()):
            return {"status_code": 400, "message": "Invalid CSV file"}

        await buffer.write(content)

    # get certificate temple path
    certificate_file = "./Data/Event_Certificate_Template.docx"

    # get participants
    list_participate = await process_csv(csv_file_path)

    await create_docx_files(
        certificate_file, list_participate, event=event_name, ambassador=ambassador_name
    )

    additional_files = [mailerpath]

    zip_folder("./Output/PDF", zip_filename, additional_files)

    return FileResponse(
        zip_filename, media_type="application/octet-stream", filename="certificates.zip"
    )


if __name__ == "__main__":
    # Run the app with Uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
